from openai import OpenAI
import openpyxl
from openpyxl.utils import get_column_letter
from prompts import PROMPT1, PROMPT2, PROMPT3, PROMPT4, PROMPT5, PROMPT6, PROMPT7, PROMPT8, PROMPT9
import random
import pandas as pd
import json

prompt = """
         from manim import *


def circle_with_text(
    r: int, txt: str, color=BLACK, font_size: int = 10, font_color=BLACK, stroke_width: int = 2, fill_color=BLACK
) -> Mobject:
    c = Circle(r, color).set_opacity(1).set_fill(fill_color)
    t = Tex(txt, stroke_width=stroke_width,
            font_size=font_size, color=font_color)
    return VGroup(c, t)


def electron() -> Mobject:
    return circle_with_text(0.15, "e$^-$", YELLOW_D, 24, BLACK, 1, YELLOW_D)


def oxygen() -> Mobject:
    list = []
    start = 0.3
    end = 0.6
    list.append(Line(start=LEFT * start, end=LEFT * end))
    list.append(Line(start=RIGHT * start, end=RIGHT * end))
    list.append(Line(start=UP * start, end=UP * end))
    list.append(Line(start=DOWN * start, end=DOWN * end))
    list = map(lambda x: x.set_opacity(0.7), list)
    return VGroup(circle_with_text(0.285, "O$^2$", GREY_B, 25, WHITE, 1), *list)


def metal() -> Mobject:
    return circle_with_text(0.285, "Co", BLUE_D, 25, WHITE, 1)


def positive() -> Mobject:
    return circle_with_text(0.25, "+", RED_A, 32, BLACK, 4, RED_A)


def negative() -> Mobject:
    return circle_with_text(0.25, "-", BLUE_A, 32, BLACK, 4, BLUE_A)


def metal_oxid(rows: int, cols: int, scale=1.5) -> Mobject:
    list = []
    for i in range(rows):
        for j in range(cols):
            if (j + i) % 2 == 0:
                list.append(oxygen().set_x(i / scale).set_y(j / scale))
            else:
                list.append(metal().set_x(i / scale).set_y(j / scale))
    return VGroup(*list).set_x(0).set_y(0)


def lithium_grid(rows, cols) -> Mobject:
    list = []
    for i in range(rows):
        for j in range(cols):
            list.append(lithium_ion().set_x(j).set_y(i))
    return VGroup(*list).set_x(0).set_y(0)

def carbon_grid(rows: int) -> Mobject:
    list = []
    for i in range(rows):
        list.append(carbon_layer().set_y(i * 1.5))
    return VGroup(*list).set_x(0).set_y(0)


def carbon(small: int = False) -> Mobject:
    return circle_with_text(
        0.285 if small == 0 else (0.225 if small == 1 else 0.25),
        "C",
        GRAY_B,
        30 if small == 0 else (24 if small == 1 else 27),
        WHITE,
        1,
    ).set_z_index(5 if small == 1 else 10)


def carbon_layer() -> Mobject:
    list = []
    positions = [
        [-1.2, 0, 0],
        [-0.45, 0.4, 0],
        [0.45, 0.4, 0],
        [1.2, 0, 0],
        [0.55, -0.3, 0],
        [-0.55, -0.3, 0],
    ]
    small = [2, 1, 1, 2, 0, 0]
    for i in range(len(positions)):
        list.append(
            Line(positions[i], positions[i - 1], color=GREY_B).set_z_index(4))
        list.append(carbon(small[i]).set_x(positions[i][0]).set_y(positions[i][1]))
    list.append(
        Polygon(*positions, color=WHITE).set_fill(WHITE, 0.075).set_z_index(-1))
    return VGroup(*list)


def lithium_ion() -> Mobject:
    return circle_with_text(0.285, "Li$^+$", RED_D, 25, WHITE, 1)


def move_along_complex_path(obj: Mobject, stop_indices, points, run_time: float, rate_func=linear):
    animations = []
    paths = []
    for i in range(len(points) - 1):
        if i not in stop_indices:
            paths.append(Line(points[i], points[i + 1]))

    for i in range(len(paths)):
        animations.append(MoveAlongPath(
            obj, paths[i], run_time=run_time, rate_func=rate_func))

    return animations
         """


# Function to append data to an existing Excel sheet
def append_to_excel(file_path, data, sheet_name="NeuralSheet"):
    # Load the workbook and the sheet
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Find the next row in the sheet
    next_row = ws.max_row + 1

    # Append data rows
    for row_data in data:
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col_num, value=value)
        next_row += 1

    # Save the workbook
    wb.save(file_path)


def get_all_examples(datasheet):
    df = pd.read_excel(datasheet)
    training_examples = df['Manim Code'].tolist()
    return training_examples


# Path to your Excel file
excel_file_path = './manim_data_synth.xlsx'


client = OpenAI()

prompts = [PROMPT1, PROMPT2, PROMPT3, PROMPT4,
           PROMPT5, PROMPT6, PROMPT7, PROMPT8, PROMPT9]


def gen_and_save_synth_data(prompt_list):
    random_prompt = random.choice(prompt_list)

    try:
        completion = client.chat.completions.create(
            model="gpt-4-1106-preview",
            messages=[
                {"role": "system", "content": "You are obedient and direct assistant who listens to exactly what the user asks. You generate accurate Python code for the ManimCE library."},
                {"role": "user", "content": "The following is some example Manim code: \n" + random_prompt +
                 "\n Use this and your knowledge of ManimCE to come up with another unique valid code sample that is only related to the general theme or topic of the example. ONLY return the code itself and nothing else. Do not format it with Markdown."}
            ]
        )

        print(completion.choices[0].message.content)

        # Append data to the Excel file
        append_to_excel(excel_file_path, [
                        [completion.choices[0].message.content]])
    except Exception:
        print(Exception)
        pass


# example_prompts = get_all_examples("./manim_neural_data.xlsx")

# SYTHENTIC_SAMPLE_COUNT = 1000
# for x in range(0, SYTHENTIC_SAMPLE_COUNT):
#     gen_and_save_synth_data(example_prompts)


def gen_input_data(output_json_path, excel_file_path):
    output_data = []
    with open(output_json_path, 'r') as file:
        output_data = json.load(file)

    for query in output_data:
        try:
            completion = client.chat.completions.create(
                model="gpt-4-1106-preview",
                messages=[
                    {"role": "system", "content": "You are obedient and direct assistant who listens to exactly what the user asks. You generate accurate detailed prompts a user could have asked based on given ManimCE code."},
                    {"role": "user", "content": "The following is the ManimCE code generated: \n" + query["text"] +
                     "\n Use this to come up with a detailed prompt a user would have asked that led to this code being generated. ONLY return the prompt the user may have asked and nothing else. Do not format it with Markdown." +
                     "\n Remember that the user does not have any knowledge about ManimCE or code. They are only prompting to generate an animation and therefore the prompt or question you generate must reflect that."}
                ]
            )

            print(completion.choices[0].message.content)

            # Append data to the Excel file
            append_to_excel(excel_file_path, [
                            [completion.choices[0].message.content]], "Generated Input")
        except Exception:
            print(Exception)
            pass


gen_input_data("./instructgen/final_combined_animation_data.json",
               "./master_finetune_datasheet.xlsx")
